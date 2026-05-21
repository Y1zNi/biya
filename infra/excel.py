"""Excel 读写."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from config import EXPORT_DIR
from core.export_schema import get_export_headers, get_sheet_name, item_to_export_cells, normalize_platform_id
from core.platforms import list_collectable_platform_ids
from core.models import CollectResultItem, ExcelSheetData
from infra.link_extract import extract_collect_links_from_cell
from infra.platform_detect import guess_link_column_index, looks_like_collect_link


def read_excel_sheet(file_path: str) -> ExcelSheetData:
  workbook = load_workbook(file_path, read_only=True, data_only=True)
  sheet = workbook.active
  rows_iter = sheet.iter_rows(values_only=True)
  headers_row = next(rows_iter, None)
  headers = [str(cell).strip() if cell is not None else '' for cell in (headers_row or [])]
  rows: List[list] = []
  for row in rows_iter:
    rows.append(list(row))
  workbook.close()
  return ExcelSheetData(headers=headers, rows=rows)


def extract_links_from_column(sheet_data: ExcelSheetData, column_index: int) -> List[str]:
  links: List[str] = []
  for row in sheet_data.rows:
    if column_index >= len(row):
      continue
    value = row[column_index]
    if value is None:
      continue
    text = str(value).strip()
    if text:
      links.append(text)
  return links


def extract_first_column_links_with_rows(file_path: str) -> List[Tuple[int, str]]:
  """读取活动表 A 列，从第 1 行起，仅返回像链接的非空单元格."""
  workbook = load_workbook(file_path, read_only=True, data_only=True)
  sheet = workbook.active
  items: List[Tuple[int, str]] = []
  for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    if not row:
      continue
    value = row[0]
    if value is None:
      continue
    text = str(value).strip()
    if not text or not looks_like_collect_link(text):
      continue
    for url in extract_collect_links_from_cell(text):
      items.append((row_index, url))
  workbook.close()
  return items


def extract_links_from_text(text: str) -> List[Tuple[int, str]]:
  """按行解析文本；行号从 1 起，从分享文案中提取短链."""
  items: List[Tuple[int, str]] = []
  for line_index, line in enumerate((text or '').splitlines(), start=1):
    value = line.strip()
    if not value or not looks_like_collect_link(value):
      continue
    for url in extract_collect_links_from_cell(value):
      items.append((line_index, url))
  return items


def filter_links_by_row_range(
  items: List[Tuple[int, str]],
  start_row: int,
  end_row: int,
) -> List[str]:
  """按 Excel 行号过滤链接；end_row=0 表示到最后一行."""
  start = max(1, start_row)
  upper = end_row if end_row > 0 else None
  links: List[str] = []
  for row_index, link in items:
    if row_index < start:
      continue
    if upper is not None and row_index > upper:
      continue
    links.append(link)
  return links


def auto_detect_link_column(headers: List[str]) -> int:
  index = guess_link_column_index(headers)
  if index >= 0:
    return index
  return 0 if headers else -1


def export_platform_results(
  items: List[CollectResultItem],
  platform_id: str,
  file_path: Optional[str] = None,
) -> Path:
  pid = normalize_platform_id(platform_id)
  EXPORT_DIR.mkdir(parents=True, exist_ok=True)
  if not file_path:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = str(EXPORT_DIR / f'采集结果_{get_sheet_name(pid)}_{timestamp}.xlsx')

  workbook = Workbook()
  sheet = workbook.active
  sheet.title = get_sheet_name(pid)[:31]
  headers = get_export_headers(pid)
  sheet.append(headers)
  for item in items:
    sheet.append(item_to_export_cells(item, pid))
  workbook.save(file_path)
  return Path(file_path)


def export_all_platform_results(
  results_by_platform: Dict[str, List[CollectResultItem]],
  file_path: Optional[str] = None,
) -> Path:
  EXPORT_DIR.mkdir(parents=True, exist_ok=True)
  if not file_path:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = str(EXPORT_DIR / f'采集结果_全部_{timestamp}.xlsx')

  workbook = Workbook()
  workbook.remove(workbook.active)

  has_sheet = False
  ordered_ids = list_collectable_platform_ids()
  extra_ids = [pid for pid in results_by_platform if pid not in ordered_ids]
  for platform_id in [*ordered_ids, *extra_ids]:
    items = results_by_platform.get(platform_id) or []
    if not items:
      continue
    pid = normalize_platform_id(platform_id)
    sheet = workbook.create_sheet(title=get_sheet_name(pid)[:31])
    headers = get_export_headers(pid)
    sheet.append(headers)
    for item in items:
      sheet.append(item_to_export_cells(item, pid))
    has_sheet = True

  if not has_sheet:
    sheet = workbook.create_sheet(title='采集结果')
    sheet.append(get_export_headers('douyin'))

  workbook.save(file_path)
  return Path(file_path)


def export_results_to_xlsx(
  items: List[CollectResultItem],
  file_path: Optional[str] = None,
) -> Path:
  """兼容旧接口：按第一条结果平台导出，或合并为单表（仅当同一平台时合理）."""
  if not items:
    return export_platform_results([], 'unknown', file_path)
  platform_id = items[0].platform_id or 'unknown'
  unique_ids = {item.platform_id or 'unknown' for item in items}
  if len(unique_ids) == 1:
    return export_platform_results(items, platform_id, file_path)

  grouped: Dict[str, List[CollectResultItem]] = {}
  for item in items:
    pid = normalize_platform_id(item.platform_id)
    grouped.setdefault(pid, []).append(item)
  return export_all_platform_results(grouped, file_path)
